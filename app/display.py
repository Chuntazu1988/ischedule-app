import io
import re
import pandas as pd
import streamlit as st

from utils.helpers import (
    clean_text, safe_html, normalize_role_label, gender_role_label,
    is_time_text, to_datetime_time, time_to_minutes, minutes_between,
    short_flight_number, flight_key, name_key, name_key_reversed,
    classify_shift, shift_length, break_label_for_employee, required_break,
    return_text_by_shift, break_deadline_before_flight,
    employee_shift_text, gap_minutes_to_next, next_task_plain_text,
)
from app.scheduler import (
    requirements_text, get_qualified_candidates_for_swap, do_swap,
    is_within_shift,
)


# =========================
# LABEL SCHEDULE (adds text annotations to result_df)
# =========================

def build_next_task_labels(result_df, employees_df):
    df = result_df.copy()

    if df.empty:
        df["טקסט עובד"] = []
        df["תפקיד נוכחי"] = []
        df["המשך אזורי"] = []
        return df

    df["טקסט עובד"] = df["עובד"].astype(str)
    df["תפקיד נוכחי"] = df["תפקיד"].apply(normalize_role_label)
    df["המשך אזורי"] = ""

    timed_df = df[df["התחלה"].astype(str).str.strip() != ""].copy()
    timed_df["_start_dt"] = pd.to_datetime(timed_df["התחלה"], format="%H:%M", errors="coerce")

    break_state = {}   # emp -> {"stage": 0/1/2, "last_m": int}
    MIN_BETWEEN_BREAKS = 4 * 60

    for idx, row in timed_df.sort_values("_start_dt").iterrows():
        emp = str(row["עובד"]).strip()

        if "❌" in emp:
            df.loc[idx, "טקסט עובד"] = emp
            df.loc[idx, "המשך אזורי"] = ""
            continue

        emp_match = employees_df[employees_df["שם"] == emp]
        if emp_match.empty:
            next_text = next_task_plain_text(timed_df, idx, emp) or "חזרה"
            df.loc[idx, "טקסט עובד"] = f"{emp} - {next_text}"
            df.loc[idx, "המשך אזורי"] = next_text
            continue

        emp_row     = emp_match.iloc[0]
        shift_text  = employee_shift_text(employees_df, emp)
        shift_suffix = f" | {shift_text}" if shift_text else ""
        shift_type  = classify_shift(emp_row)

        next_plain    = next_task_plain_text(timed_df, idx, emp)
        task_end_dt   = to_datetime_time(row["סיום"])
        task_start_dt = to_datetime_time(row["התחלה"])
        task_end_m    = task_end_dt.hour * 60 + task_end_dt.minute
        gap_to_next   = gap_minutes_to_next(timed_df, idx, emp)
        break_label   = break_label_for_employee(emp_row)
        base_role     = normalize_role_label(str(row.get("תפקיד בסיס", "")))
        is_agent      = base_role in {"דיילת", "דייל", "שומר TSA"}

        state  = break_state.get(emp, {"stage": 0, "last_m": -9999})
        stage  = state["stage"]
        last_m = state["last_m"]

        te = task_end_m
        if last_m > 0 and te < last_m:
            te += 1440
        time_since_last = (te - last_m) if last_m != -9999 else 9999

        is_last_task = (next_plain == "")
        action = None

        def enough_time_from_shift_start():
            ss = clean_text(emp_row.get("תחילת משמרת", ""))
            if not is_time_text(ss):
                return True
            s = time_to_minutes(ss)
            te_local = task_end_m
            if te_local < s: te_local += 1440
            return (te_local - s) >= MIN_BETWEEN_BREAKS

        long_shift    = shift_length(emp_row) > 9 * 60
        can_break_now = (not long_shift) or enough_time_from_shift_start()

        if break_label == "הפסקה ורענון":
            if stage == 0 and can_break_now:
                action = "break"
            elif stage == 1 and time_since_last >= MIN_BETWEEN_BREAKS:
                action = "refresh"
        elif break_label in ("הפסקה", "רענון"):
            if stage == 0 and can_break_now:
                action = break_label

        if shift_type == "night" and stage == 0 and break_label and can_break_now:
            action = "break" if "הפסקה" in break_label else "רענון"

        if action is None and is_agent and gap_to_next is not None and gap_to_next >= 30:
            if stage == 0 and break_label and can_break_now:
                action = "break" if "הפסקה" in break_label else "רענון"

        def do_action(act, continuation):
            nonlocal stage
            label = "הפסקה" if act == "break" else "רענון" if act == "refresh" else act
            if act in ("break", "הפסקה"):
                break_state[emp] = {"stage": 1 if break_label == "הפסקה ורענון" else 2, "last_m": task_end_m}
            elif act in ("refresh", "רענון"):
                break_state[emp] = {"stage": 2, "last_m": task_end_m}
            return f"{label} ו{continuation}" if continuation else label

        return_text = return_text_by_shift(emp_row, task_end_dt)

        if is_last_task:
            if break_label == "הפסקה ורענון" and stage == 1 and time_since_last >= MIN_BETWEEN_BREAKS:
                action = "refresh"
            continuation = return_text
        else:
            continuation = next_plain

        if action:
            next_text = do_action(action, continuation)
        else:
            next_text = continuation if continuation else return_text

        task_start_m   = task_start_dt.hour * 60 + task_start_dt.minute
        break_deadline = break_deadline_before_flight(emp_row, task_start_m)
        deadline_suffix = (
            f" [הפסקה עד {break_deadline}]"
            if break_deadline and stage == 0
            else ""
        )

        df.loc[idx, "טקסט עובד"] = f"{emp} - {next_text}{shift_suffix}{deadline_suffix}"
        df.loc[idx, "המשך אזורי"] = next_text

    return df


# =========================
# OUTPUT TABLES
# =========================

def build_output_table(flights_df, result_labeled, employees_df):
    rows = []
    flights_df = flights_df.copy()
    flights_df["_flight_key"] = flights_df["טיסה"].apply(flight_key)
    flights_df = flights_df.drop_duplicates(subset=["_flight_key"], keep="first").drop(columns=["_flight_key"])

    for _, flight in flights_df.iterrows():
        fnum     = str(flight["טיסה"]).strip()
        dep      = clean_text(flight.get("המראה", ""))
        boarding = clean_text(flight.get("בורדינג", ""))
        aircraft = clean_text(flight.get("סוג מטוס", ""))
        reg      = clean_text(flight.get("רישוי", ""))
        reqs     = requirements_text(flight)

        tasks = result_labeled[result_labeled["טיסה"].astype(str).str.strip() == fnum]

        management_lines = []
        agents_lines     = []
        management_roles = []
        agent_roles      = []

        for _, task in tasks.iterrows():
            role       = str(task["תפקיד"])
            base_role  = normalize_role_label(role)
            worker     = str(task.get("עובד", ""))
            text_value = str(task["טקסט עובד"])

            display_role = gender_role_label(role, employees_df, worker)

            if "❌" not in worker and worker and " - " in text_value:
                name_part, rest = text_value.split(" - ", 1)
                text_value = f"{name_part} ({display_role}) - {rest}"
            elif "❌" not in worker and worker:
                text_value = f"{worker} ({display_role})"

            shift = employee_shift_text(employees_df, worker)
            if shift and "❌" not in worker and f"({shift})" not in text_value:
                text_value = f"{text_value} ({shift})"

            item = {"text": text_value, "role": base_role}

            if "ראש צוות" in role or "טרייני" in role or "מתאם" in role or "מפקח" in role:
                management_lines.append(item)
                if base_role not in management_roles:
                    management_roles.append(base_role)
            elif "דייל" in role or "שומר" in role or "בדיקת טרייני" in role:
                agents_lines.append(item)
                if base_role not in agent_roles:
                    agent_roles.append(base_role)

        rows.append({
            "מספר טיסה": fnum,
            "יעד": clean_text(flight["יעד"]),
            "זמנים": f"{dep} ({boarding})" if boarding else dep,
            "מטוס/רישוי": f"{aircraft}\n{reg}".strip(),
            "תפקידים דרושים": " | ".join(reqs),
            "כותרת ניהול": " / ".join(management_roles) if management_roles else "ניהול",
            "כותרת דיילים": " / ".join(agent_roles) if agent_roles else "דיילים",
            "דיילים / שומר TSA": "\n".join([f"{x['role']}||{x['text']}" for x in agents_lines]),
            "ראש צוות / מתאם תורים / מפקח TSA": "\n".join([f"{x['role']}||{x['text']}" for x in management_lines]),
        })

    return pd.DataFrame(rows)


def build_workload(result_df, employees_df):
    rows = []

    if result_df.empty:
        return pd.DataFrame(columns=["עובד", "משימות", "דקות עבודה", "הפסקה נדרשת", "סה״כ כולל הפסקות"])

    timed = result_df[result_df["התחלה"].astype(str).str.strip() != ""].copy()

    # בנה מפת נורמליזציה: name_key(וריאציה) → שם קנוני מ-employees_df
    # כך "שחר פרגסליך" ו"פרגסליך שחר" יאוחדו לאותו שם
    name_canon: dict = {}
    for _, emp in employees_df.iterrows():
        canon = clean_text(emp.get("שם", ""))
        if not canon:
            continue
        name_canon[name_key(canon)] = canon
        name_canon[name_key_reversed(canon)] = canon

    def canonicalize(w):
        w = clean_text(str(w))
        if "❌" in w:
            return w
        return name_canon.get(name_key(w), w)

    timed["עובד"] = timed["עובד"].apply(canonicalize)

    real_workers = sorted([
        worker for worker in timed["עובד"].dropna().unique()
        if "❌" not in str(worker)
    ])

    for emp in real_workers:
        tasks = timed[timed["עובד"] == emp]
        total = sum(
            minutes_between(to_datetime_time(task["התחלה"]), to_datetime_time(task["סיום"]))
            for _, task in tasks.iterrows()
        )
        emp_row  = employees_df[employees_df["שם"] == emp]
        break_min = required_break(emp_row.iloc[0]) if not emp_row.empty else 0
        rows.append({
            "עובד":               emp,
            "משימות":             len(tasks),
            "דקות עבודה":         total,
            "הפסקה נדרשת":       break_min,
            "סה״כ כולל הפסקות":  total + break_min,
        })

    return pd.DataFrame(rows)


def build_counter_continuity_rows(result_labeled, employees_df):
    rows = []

    if result_labeled.empty:
        return pd.DataFrame(columns=["עובד", "משמרת", "טיסות", "טיסות משובצות", "תפקידים", "הערה"])

    for _, emp in employees_df.iterrows():
        name  = emp["שם"]
        shift = employee_shift_text(employees_df, name)

        emp_tasks = result_labeled[
            (result_labeled["עובד"].astype(str) == name) &
            (result_labeled["התחלה"].astype(str).str.strip() != "")
        ].copy()

        if emp_tasks.empty:
            rows.append({
                "עובד": name, "משמרת": shift, "טיסות": 0,
                "טיסות משובצות": "", "תפקידים": "",
                "הערה": "לא שובץ לטיסות. נשאר בדלפקים או לפי הנחיית אחמ״ש.",
            })
            continue

        rows.append({
            "עובד":          name,
            "משמרת":         shift,
            "טיסות":         len(emp_tasks),
            "טיסות משובצות": " | ".join(emp_tasks["טיסה"].astype(str).tolist()),
            "תפקידים":       " | ".join(emp_tasks["תפקיד"].astype(str).tolist()),
            "הערה":          "רצף טיסות" if len(emp_tasks) >= 2 else "טיסה בודדת ואז חזרה לדלפקים/סיום משמרת",
        })

    return pd.DataFrame(rows)


def build_available_in_hall(schedule_df, employees_df, flights_df):
    timed = schedule_df[
        (schedule_df["התחלה"].astype(str).str.strip() != "") &
        (~schedule_df["עובד"].astype(str).str.contains("❌", na=False))
    ].copy()

    if timed.empty:
        return pd.DataFrame(columns=["עובד", "תפקיד עיקרי", "משמרת", "פנוי מ", "פנוי עד", "פנות (דק׳)", "משימה הבאה", "הערה"])

    timed["_start_dt"] = pd.to_datetime(timed["התחלה"], format="%H:%M", errors="coerce")
    timed["_end_dt"]   = pd.to_datetime(timed["סיום"],   format="%H:%M", errors="coerce")

    hall_workers = timed["עובד"].unique()
    rows = []

    for emp_name in hall_workers:
        emp_tasks  = timed[timed["עובד"] == emp_name].sort_values("_start_dt")
        emp_row_df = employees_df[employees_df["שם"] == emp_name]
        emp_row    = emp_row_df.iloc[0] if not emp_row_df.empty else None

        shift_start_str = clean_text(emp_row.get("תחילת משמרת", "")) if emp_row is not None else ""
        shift_end_str   = clean_text(emp_row.get("סוף משמרת",   "")) if emp_row is not None else ""

        if not is_time_text(shift_start_str) or not is_time_text(shift_end_str):
            continue

        shift_text = f"{shift_start_str}-{shift_end_str}"
        main_role  = emp_tasks["תפקיד בסיס"].mode().iloc[0] if not emp_tasks.empty else ""

        try:
            shift_start_dt = pd.to_datetime(shift_start_str, format="%H:%M")
            shift_end_dt   = pd.to_datetime(shift_end_str,   format="%H:%M")
            if shift_end_dt <= shift_start_dt:
                shift_end_dt += pd.Timedelta(hours=24)
        except Exception:
            continue

        task_list = emp_tasks.reset_index(drop=True)
        gaps = []

        for i in range(len(task_list) - 1):
            end_i      = task_list.loc[i,   "_end_dt"]
            start_next = task_list.loc[i+1, "_start_dt"]
            if start_next < end_i:
                start_next += pd.Timedelta(hours=24)
            gap_min = int((start_next - end_i).total_seconds() / 60)
            if gap_min < 20 or gap_min > 300:
                continue
            next_task_text = f"{task_list.loc[i+1, 'טיסה']} — {normalize_role_label(task_list.loc[i+1, 'תפקיד'])}"
            gaps.append({
                "from": end_i.strftime("%H:%M"),
                "to":   start_next.strftime("%H:%M"),
                "gap":  gap_min,
                "next": next_task_text,
                "note": "פנוי בין טיסות",
            })

        if not task_list.empty:
            last_end = task_list.iloc[-1]["_end_dt"]
            le = last_end
            if le < shift_start_dt:
                le += pd.Timedelta(hours=24)
            se = shift_end_dt
            if se < le:
                se += pd.Timedelta(hours=24)
            remaining = int((se - le).total_seconds() / 60)
            if 30 <= remaining <= 300:
                gaps.append({
                    "from": last_end.strftime("%H:%M"),
                    "to":   shift_end_str,
                    "gap":  remaining,
                    "next": "סיום משמרת",
                    "note": "פנוי לאחר סיום טיסות",
                })

        for gap in gaps:
            rows.append({
                "עובד":         emp_name,
                "תפקיד עיקרי": normalize_role_label(main_role),
                "משמרת":        shift_text,
                "פנוי מ":       gap["from"],
                "פנוי עד":      gap["to"],
                "פנות (דק׳)":  gap["gap"],
                "משימה הבאה":   gap["next"],
                "הערה":         gap["note"],
            })

    if not rows:
        return pd.DataFrame(columns=["עובד", "תפקיד עיקרי", "משמרת", "פנוי מ", "פנוי עד", "פנות (דק׳)", "משימה הבאה", "הערה"])

    return pd.DataFrame(rows).sort_values(["פנוי מ", "עובד"]).reset_index(drop=True)


def build_unassigned_agents(schedule_df, employees_df, shift_map):
    assigned_workers = set(
        schedule_df[~schedule_df["עובד"].astype(str).str.contains("❌", na=False)]["עובד"].tolist()
    )
    rows = []
    for _, emp in employees_df.iterrows():
        name = clean_text(emp.get("שם", ""))
        if not name:
            continue
        shift_start = clean_text(emp.get("תחילת משמרת", ""))
        shift_end   = clean_text(emp.get("סוף משמרת", ""))
        if not is_time_text(shift_start) or not is_time_text(shift_end):
            continue
        if name in assigned_workers:
            continue
        is_agent = str(emp.get("דייל", "")).strip() == "כן"
        is_tl    = str(emp.get("ראש צוות", "")).strip() == "כן"
        if not is_agent and not is_tl:
            continue
        role = "ראש צוות" if is_tl else "דייל"
        rows.append({
            "שם":           name,
            "תפקיד":        role,
            "משמרת":        f"{shift_start}-{shift_end}",
            "תחילת משמרת": shift_start,
            "סוף משמרת":   shift_end,
        })
    if not rows:
        return pd.DataFrame(columns=["שם", "תפקיד", "משמרת", "תחילת משמרת", "סוף משמרת"])
    df = pd.DataFrame(rows)
    return df.sort_values(["תחילת משמרת", "שם"]).reset_index(drop=True)


# =========================
# RENDERING FUNCTIONS
# =========================

def line_style_by_role(current_role, line):
    if "❌" in str(line):
        return "role-missing"
    role = normalize_role_label(current_role)
    if role == "ראש צוות":   return "role-teamlead"
    if role == "מפקח TSA":   return "role-inspector"
    if role == "שומר TSA":   return "role-guard"
    if role == "טרייני ר״צ": return "role-trainee"
    if role == "מתאם תורים": return "role-queue"
    if role == "דיילת":      return "role-agent"
    return "role-agent"


def render_line(line, current_role=""):
    line_str   = str(line)
    style_cls  = line_style_by_role(current_role, line_str)
    shift_badge = ""
    shift_match = re.search(r"\((\d{2}:\d{2}-\d{2}:\d{2})\)\s*$", line_str)
    if shift_match:
        shift_badge = shift_match.group(1)
        line_str = line_str[:shift_match.start()].rstrip()

    badge_html = (
        f'<span style="float:left;background:#e8f0fe;color:#1a3d7a;'
        f'font-size:11px;font-weight:900;border-radius:6px;'
        f'padding:2px 7px;margin-right:6px;white-space:nowrap;">🕐 {safe_html(shift_badge)}</span>'
        if shift_badge else ""
    )
    st.markdown(
        f'<div class="assignment-line {style_cls}">'
        f'{badge_html}'
        f'{safe_html(line_str)}'
        f'</div>',
        unsafe_allow_html=True,
    )


def render_flight_card(row):
    aircraft   = str(row["מטוס/רישוי"]).replace("\n", " / ")
    reqs       = str(row["תפקידים דרושים"])
    left_text  = str(row["ראש צוות / מתאם תורים / מפקח TSA"])
    right_text = str(row["דיילים / שומר TSA"])

    required_line = " | ".join([part.strip() for part in reqs.split("|") if part.strip()]) or "לא הוגדרו תפקידים"

    st.markdown(
        f"""
        <div class="flight-card">
          <div class="flight-head">
            <div class="flight-row">
              <div class="flight-name">✈️ {safe_html(row['מספר טיסה'])} ← {safe_html(row['יעד'])}</div>
              <div class="flight-meta">🕒 {safe_html(row['זמנים'])} | 🛩️ {safe_html(aircraft)}</div>
            </div>
            <div class="req-line">תפקידים דרושים: {safe_html(required_line)}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    management_title = str(row.get("כותרת ניהול", "ניהול"))
    st.markdown(f'<div class="panel-title">👔 {safe_html(management_title)}</div>', unsafe_allow_html=True)
    if left_text and left_text != "nan":
        for line in left_text.split("\n"):
            if "||" in line:
                role_part, text_part = line.split("||", 1)
                render_line(text_part, role_part)
            else:
                render_line(line)
    else:
        render_line("אין שיבוץ")

    st.markdown("<div style='margin-top:6px'></div>", unsafe_allow_html=True)

    agents_title = str(row.get("כותרת דיילים", "דיילים"))
    st.markdown(f'<div class="panel-title">🧍 {safe_html(agents_title)}</div>', unsafe_allow_html=True)
    if right_text and right_text != "nan":
        for line in right_text.split("\n"):
            if "||" in line:
                role_part, text_part = line.split("||", 1)
                render_line(text_part, role_part)
            else:
                render_line(line)
    else:
        render_line("אין שיבוץ")


def render_flight_card_with_swap(row, schedule_df, employees_df):
    aircraft   = str(row["מטוס/רישוי"]).replace("\n", " / ")
    reqs       = str(row["תפקידים דרושים"])
    left_text  = str(row["ראש צוות / מתאם תורים / מפקח TSA"])
    right_text = str(row["דיילים / שומר TSA"])
    fnum       = str(row["מספר טיסה"]).strip()

    required_line  = " | ".join([p.strip() for p in reqs.split("|") if p.strip()]) or "לא הוגדרו תפקידים"
    aircraft_short = str(row["מטוס/רישוי"]).split("\n")[0].strip()

    has_missing  = "❌" in left_text or "❌" in right_text
    missing_icon = " ⚠️" if has_missing else ""

    expander_label = (
        f"✈️ {fnum} ← {row['יעד']}{missing_icon}"
        f"   |   🕒 {row['זמנים']}"
        f"   |   🛩️ {aircraft_short}"
        f"   |   {required_line}"
    )

    with st.expander(expander_label, expanded=False):

        def render_lines_with_swap(panel_lines):
            for line_i, line in enumerate(panel_lines):
                role_part, text_part = (line.split("||", 1) if "||" in line else ("", line))

                worker_raw = text_part.split(" - ")[0].strip() if " - " in text_part else text_part.split(" (")[0].strip()
                role_base  = normalize_role_label(role_part) if role_part else ""

                flight_tasks = schedule_df[schedule_df["טיסה"].astype(str).str.strip() == fnum]
                match = flight_tasks[
                    (flight_tasks["עובד"].astype(str).str.contains(re.escape(worker_raw), na=False)) &
                    (flight_tasks["תפקיד בסיס"].astype(str).apply(normalize_role_label) == role_base)
                ]
                if match.empty:
                    match = flight_tasks[
                        flight_tasks["תפקיד בסיס"].astype(str).apply(normalize_role_label) == role_base
                    ]

                if match.empty:
                    render_line(text_part, role_part)
                    continue

                task_idx  = match.index[0]
                uid       = f"{fnum}_{task_idx}_{line_i}"
                popup_key = f"popup_open_{uid}"
                if popup_key not in st.session_state:
                    st.session_state[popup_key] = False

                style_cls   = line_style_by_role(role_part, text_part)
                shift_badge = ""
                shift_match = re.search(r"\((\d{2}:\d{2}-\d{2}:\d{2})\)\s*$", text_part)
                if shift_match:
                    shift_badge  = shift_match.group(1)
                    display_text = text_part[:shift_match.start()].rstrip()
                else:
                    display_text = text_part

                badge_html = (
                    f'<span style="float:left;background:#e8f0fe;color:#1a3d7a;'
                    f'font-size:11px;font-weight:900;border-radius:6px;'
                    f'padding:2px 7px;margin-right:6px;white-space:nowrap;">🕐 {safe_html(shift_badge)}</span>'
                    if shift_badge else ""
                )

                col_line, col_arrow = st.columns([11, 1])
                with col_line:
                    st.markdown(
                        f'<div class="assignment-line {style_cls}" style="margin-bottom:0">'
                        f'{badge_html}{safe_html(display_text)}</div>',
                        unsafe_allow_html=True,
                    )
                with col_arrow:
                    arrow = "▲" if st.session_state[popup_key] else "▼"
                    if st.button(arrow, key=f"arrow_{uid}", help="החלף עובד"):
                        st.session_state[popup_key] = not st.session_state[popup_key]
                        st.rerun()

                if st.session_state[popup_key]:
                    candidates = get_qualified_candidates_for_swap(
                        schedule_df, employees_df, fnum, role_base, task_idx
                    )
                    st.markdown(
                        f'<div class="swap-popup">'
                        f'<div class="swap-popup-title">🔄 החלפת עובד — {safe_html(worker_raw)} ({safe_html(role_base)})</div>',
                        unsafe_allow_html=True,
                    )

                    if not candidates:
                        st.warning("אין עובדים מוסמכים ופנויים להחלפה כרגע.")
                    else:
                        st.markdown('<div class="swap-popup-label">בחר עובד חלופי:</div>', unsafe_allow_html=True)
                        selected_new = st.selectbox("", options=candidates, key=f"swap_select_{uid}", label_visibility="collapsed")
                        st.markdown(f'<div class="swap-popup-label">מה לעשות עם {safe_html(worker_raw)}?</div>', unsafe_allow_html=True)
                        action = st.radio("", options=["השאר ללא שיבוץ", "העבר לחריץ פנוי בטיסה אחרת"],
                                          key=f"swap_action_{uid}", horizontal=True, label_visibility="collapsed")

                        target_flight = None
                        if action == "העבר לחריץ פנוי בטיסה אחרת":
                            other_flights = sorted(
                                schedule_df[
                                    (schedule_df["טיסה"].astype(str).str.strip() != fnum) &
                                    (schedule_df["עובד"].astype(str).str.contains("❌")) &
                                    (schedule_df["תפקיד בסיס"].astype(str).apply(normalize_role_label) == role_base)
                                ]["טיסה"].astype(str).str.strip().unique().tolist()
                            )
                            if not other_flights:
                                st.info("אין חריצים פנויים מתאימים בטיסות אחרות.")
                            else:
                                st.markdown('<div class="swap-popup-label">בחר טיסה יעד:</div>', unsafe_allow_html=True)
                                target_flight = st.selectbox("", options=other_flights, key=f"swap_target_{uid}", label_visibility="collapsed")

                        bc1, bc2 = st.columns(2)
                        with bc1:
                            if st.button("✅ אשר החלפה", key=f"swap_confirm_{uid}", use_container_width=True):
                                displaced_action = "move" if action == "העבר לחריץ פנוי בטיסה אחרת" else "unassign"
                                updated = do_swap(st.session_state["schedule_df"], task_idx, selected_new, displaced_action, target_flight)
                                st.session_state["schedule_df"] = updated
                                st.session_state[popup_key] = False
                                st.rerun()
                        with bc2:
                            if st.button("✖ ביטול", key=f"swap_cancel_{uid}", use_container_width=True):
                                st.session_state[popup_key] = False
                                st.rerun()

                    st.markdown('</div>', unsafe_allow_html=True)
                st.markdown("<div style='margin-bottom:7px'></div>", unsafe_allow_html=True)

        management_title = str(row.get("כותרת ניהול", "ניהול"))
        agents_title     = str(row.get("כותרת דיילים", "דיילים"))

        st.markdown(f'<div class="panel-title">👔 {safe_html(management_title)}</div>', unsafe_allow_html=True)
        left_lines = [l for l in left_text.split("\n") if l.strip()] if left_text and left_text != "nan" else []
        if left_lines:
            render_lines_with_swap(left_lines)
        else:
            render_line("אין שיבוץ")

        st.markdown("<div style='margin-top:6px'></div>", unsafe_allow_html=True)

        st.markdown(f'<div class="panel-title">🧍 {safe_html(agents_title)}</div>', unsafe_allow_html=True)
        right_lines = [l for l in right_text.split("\n") if l.strip()] if right_text and right_text != "nan" else []
        if right_lines:
            render_lines_with_swap(right_lines)
        else:
            render_line("אין שיבוץ")


# =========================
# EXCEL EXPORT
# =========================

def to_excel_bytes(output_df, workload_df, schedule_df, continuity_df=None):
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            output_df.to_excel(writer, index=False, sheet_name="שיבוץ")
            workload_df.to_excel(writer, index=False, sheet_name="עומס")
            schedule_df.to_excel(writer, index=False, sheet_name="פירוט גולמי")
            if continuity_df is not None:
                continuity_df.to_excel(writer, index=False, sheet_name="רצף אזורי")
    except Exception:
        output = io.BytesIO()
        output_df.to_csv(output, index=False, encoding="utf-8-sig")
    output.seek(0)
    return output
    from io import BytesIO
    from openpyxl import load_workbook
    from io import BytesIO
    from openpyxl import load_workbook


def to_departures_report_excel_bytes(flights_df, schedule_df, employees_df):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "דוח שיבוץ טיסות - המראות"

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A4"

    headers = [
        "טיסה",
        "שעה",
        "יעד",
        "מטוס",
        "ר״צ",
        "משמרת",
        "דיילים",
        "משמרת",
    ]

    widths = [14, 14, 18, 16, 32, 18, 38, 18]

    blue_fill = PatternFill("solid", fgColor="4FB3D8")
    title_font = Font(name="Arial", size=18, bold=True)
    header_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "שיבוץ טיסות בוקר"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    ws.merge_cells("A2:H2")
    ws["A2"] = ""
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = blue_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2, wrap_text=True)

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    current_row = 4

    for _, flight in flights_df.iterrows():
        flight_num = str(flight.get("טיסה", "")).strip()
        dep_time = str(flight.get("המראה", "")).strip()
        destination = str(flight.get("יעד", "")).strip()
        aircraft = str(flight.get("סוג מטוס", "")).strip() or str(flight.get("מטוס", "")).strip()

        tasks = schedule_df[schedule_df["טיסה"].astype(str).str.strip() == flight_num].copy()

        managers = []
        agents = []

        for _, task in tasks.iterrows():
            worker = str(task.get("עובד", "")).strip()
            role = str(task.get("תפקיד בסיס", "")).strip()

            if not worker:
                continue

            try:
                shift = employee_shift_text(employees_df, worker)
            except Exception:
                shift = ""

            text = f"{worker} ({role})" if role else worker
            item = {"shift": shift, "text": text}

            if (
                "ראש צוות" in role
                or "רצ" in role
                or "ר״צ" in role
                or "TSA" in role
                or "מתאם" in role
                or "מפקד" in role
                or "טרייני" in role
            ):
                managers.append(item)
            else:
                agents.append(item)

        max_rows = max(len(managers), len(agents), 1)

        for i in range(max_rows):
            values = [
                flight_num if i == 0 else "",
                dep_time if i == 0 else "",
                destination if i == 0 else "",
                aircraft if i == 0 else "",
                managers[i]["text"] if i < len(managers) else "",
                managers[i]["shift"] if i < len(managers) else "",
                agents[i]["text"] if i < len(agents) else "",
                agents[i]["shift"] if i < len(agents) else "",
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                    readingOrder=2,
                    wrap_text=True,
                )

            current_row += 1

        current_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()